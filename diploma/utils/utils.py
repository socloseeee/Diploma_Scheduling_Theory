import json
import numpy as np

from pathlib import Path
from copy import deepcopy
from openpyxl import Workbook
from matplotlib import pyplot as plt
from PyQt5.QtGui import QPixmap, QImage
from typing import TextIO, Optional, Union

from diploma.Data import DataApp


def writing_in_files(
        result_file: TextIO = None,
        work_time: list[str] = None,
        results: list[str] = None,
) -> None:
    for elapsed_time, result in zip(work_time, results):
        print(
            f"\nElapsed time | Время работы: {elapsed_time}\n"
            f"Result | Результат: {result}"
        )
        result_file.write(
            f"\nElapsed time | Время работы: {elapsed_time[5:-4]}\n"
            f"Result | Результат: {result[5:-4]}\n"
        )


def json_open(namefile: Optional[Union[str, Path]], write_method: str, data: dict = None) -> None:
    if write_method == 'r':
        with open(namefile, 'r', encoding='utf-8') as file:
            return json.load(file)
    new_data = deepcopy(data)
    with open(namefile, 'r', encoding='utf-8') as file:
        try:
            json_dict = json.load(file)
            json_dict.update(new_data)
        except Exception as e:
            print(e)
    with open(namefile, write_method, encoding='utf-8') as file:
        json.dump(json_dict, file, indent=4, ensure_ascii=False)


def qpixmap_matrix(label, data):
    data = DataApp()
    m, n = data.data["m"], data.data["n"]
    matrix = np.fromstring(data.data["matrix"], sep=' ', dtype=int).reshape(m, n)
    matrix = np.concatenate([matrix, matrix, matrix])

    fig, ax = plt.subplots(figsize=(6, 7.4))

    ax.axis('tight')
    ax.axis('off')
    widths = [0.1 / max([len(str(val)) for val in row]) for row in matrix]
    table = ax.table(
        colWidths=widths,
        cellLoc="center",
        cellText=matrix.tolist(),
        cellColours=[['black'] * n] * m * 3,
        loc='center'
    )
    # table.scale(1, 4)
    fig.set_facecolor('black')

    plt.tight_layout()
    for cell in table.get_celld().values():
        cell.set_text_props(color='white')

    # ax.set_xlim(0, 1)
    # ax.set_ylim(0, 1)

    fig.canvas.draw()
    buf = fig.canvas.buffer_rgba()
    qimage = QImage(buf, buf.shape[1], buf.shape[0], QImage.Format_RGBA8888)
    pixmap = QPixmap(qimage)
    return pixmap


def save_table_to_xlsx(table, file_path):
    wb = Workbook()
    ws = wb.active

    # Запись заголовков
    for col in range(table.columnCount()):
        header = table.horizontalHeaderItem(col).text()
        style = None
        if header in ('Разбиение', 'Границы', 'Сортировка'):
            style = 'Accent1'
        if style is not None:
            ws.cell(row=1, column=col + 1, value=header).style = style
        else:
            ws.cell(row=1, column=col + 1, value=header)

    # Запись данных
    for row in range(1, table.rowCount() + 1):
        for col in range(table.columnCount()):
            item = table.item(row - 1, col)
            if item is not None:
                ws.cell(row=row + 1, column=col + 1, value=item.text())

    wb.save(file_path)


def save_all_tables_to_xlsx(tables, file_prefix):
    for i, table in enumerate(tables):
        file_path = f"xlsx/{file_prefix}_{i + 1}.xlsx"
        save_table_to_xlsx(table, file_path)
